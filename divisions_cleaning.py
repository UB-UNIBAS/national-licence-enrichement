from openpyxl import load_workbook, Workbook
import re
import os


university_basel = re.compile('(universit(y|ies) (of )?bas(el|le)|bas(el|le) university)'
                              '|(universit([äa]|ae)t basel'
                              '|basel universit([äa]|ae)t)', re.IGNORECASE)

university_hospital_basel = re.compile('(universit([äa]|ae)tsspital basel)'
                                       '|(university women\'s clinic basel)'
                                       '|(university (children\'s )?(hospital[s]?'
                                       '|clinic[s]?)[,]? (of )?basel)'
                                       '|(basel university hospital)'
                                       '|(university-hosp\. basel)'
                                       '|(((university hospital)|universitätsklinik(en)?|universitiitsklinik).+basel)'
                                       '|(university hospital, basel)', re.IGNORECASE)

canton_hospital_basel = re.compile('(canton hospital[,]? basel)'
                                   '|(kantonsspital basel)'
                                   '|kantonsspital[s]?.+basel', re.IGNORECASE)

biozentrum = re.compile('biozentrum.+basel', re.IGNORECASE)
institute_of_botany = re.compile('institute of botany.+basel', re.IGNORECASE)
friedrich_miescher = re.compile('friedrich[\- ]miescher[ \-]institut[e]?', re.IGNORECASE)


sti = re.compile('(swiss tropical (and public health )?institute)|(swiss tph)', re.IGNORECASE)

email = re.compile('@unibas\.ch', re.IGNORECASE)
basel_address = re.compile('bernouuianum'
                           '|schönbeinstr(\.|asse)'
                           '|rheinsprung[ ]?9'
                           '|petersgraben 9'
                           '|nadelberg 6', re.IGNORECASE)

university_hospital_not_in_basel = re.compile('university hospital', re.IGNORECASE)
private_industry = re.compile('novartis|ciba-geigy|ciba|geigy|sandoz|'
                              'roche |hoffmann[\- ]la[ ]?roche|actelion|'
                              'basel institute for immunology|syngenta|'
                              'healthecon ag, basel|basilea pharma|center for outcomes research', re.IGNORECASE)

unaffiliated_institutes = re.compile('(basel university medical clinic)'
                                     '|(zürich-basel)'
                                     '|(swiss institute of bioinformatics)', re.IGNORECASE)
other_universities = re.compile('university of zurich'
                                '|université de lausanne'
                                '|rockefeller university'
                                '|university of california', re.IGNORECASE)

fachhochschule_basel = re.compile('university of applied sciences basel', re.IGNORECASE)

work_book = load_workbook('unibas.xlsx')
sheet = work_book.active

output = Workbook()

sheets_names = ['uni-basel', 'unispital-basel', 'kantons-spital-basel', 'biozentrum', 'friedrich-miescher',
                'institute-of-botany', 'swiss-tropical-institute', 'address-in-basel', 'email', 'unispital-not-in-basel',
                'private-industry', 'unaffiliated-institutes', 'other-unis', 'not-in-basel', 'fachhochschule-basel',
                'other']

for name in sheets_names:
    output.create_sheet(name)
    output[name].append([cell.value for cell in sheet[1]])


def check_affiliations(all_affiliations, regex, file_name, relevant_row):
    for affil in all_affiliations:
        if regex.search(affil):
            values = [cell.value for cell in relevant_row]
            values.append(affil)
            output[file_name].append(values)
            with open('output/' + file_name + '.csv', 'a', encoding='utf-8') as csvfile:
                for v in values:
                    v = str(v).strip('"')
                    if v != 'None':
                        csvfile.write('"' + str(v) + '",')
                    else:
                        csvfile.write('"",')
                csvfile.write('\n')
            return True

    return False


for root, dirs, files in os.walk('output/'):
    for file in files:
        os.remove(root + file)

for row in sheet.iter_rows(min_row=2, max_col=28):
    row_affiliations = row[11].value.split(';')

    if check_affiliations(row_affiliations, university_basel, 'uni-basel', row):
        continue
    if check_affiliations(row_affiliations, university_hospital_basel, 'unispital-basel', row):
        continue
    if check_affiliations(row_affiliations, canton_hospital_basel, 'kantons-spital-basel', row):
        continue
    if check_affiliations(row_affiliations, biozentrum, 'biozentrum', row):
        continue
    if check_affiliations(row_affiliations, friedrich_miescher, 'friedrich-miescher', row):
        continue
    if check_affiliations(row_affiliations, institute_of_botany, 'institute-of-botany', row):
        continue
    if check_affiliations(row_affiliations, sti, 'swiss-tropical-institute', row):
        continue
    if check_affiliations(row_affiliations, basel_address, 'address-in-basel', row):
        continue
    if check_affiliations(row_affiliations, university_hospital_not_in_basel, 'unispital-not-in-basel', row):
        continue
    if check_affiliations(row_affiliations, private_industry, 'private-industry', row):
        continue
    if check_affiliations(row_affiliations, unaffiliated_institutes, 'unaffiliated-institutes', row):
        continue
    if check_affiliations(row_affiliations, other_universities, 'other-unis', row):
        continue
    if check_affiliations(row_affiliations, fachhochschule_basel, 'fachhochschule-basel', row):
        continue
    if check_affiliations(row_affiliations, email, 'email', row):
        continue

    cell_values = [cell.value for cell in row]
    output['other'].append(cell_values)
    # only gets here if no other search matches.
    with open('output/other.csv', 'a', encoding='utf-8') as csvfile:
        for val in cell_values:
            val = str(val).strip('"')
            if val != 'None':
                csvfile.write('"' + str(val) + '",')
            else:
                csvfile.write('"",')
        csvfile.write('\n')

output.save('output/sorted_publications.xlsx')

